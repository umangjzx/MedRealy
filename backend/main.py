"""
MedRelay — FastAPI Backend
- JWT authentication with bcrypt password hashing
- Role-based access control (admin / nurse)
- Security headers, rate limiting, request logging
- WebSocket endpoint for real-time handoff pipeline
- REST demo endpoint
- Sessions CRUD (SQLite via aiosqlite)
"""

import hashlib
import json
import traceback
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Optional

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException, UploadFile, File, Depends, Request, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, Response
from pydantic import BaseModel, Field as PydanticField, field_validator
from openpyxl import load_workbook

from backend.pipeline import HandoffPipeline
from backend.agents.relay_agent import RelayAgent, transcribe_buffer
from backend.agents.sentinel_agent import SentinelAgent
from backend.agents.staffing_agent import StaffingAgent
from backend.agents.cmio_agent import CMIOAgent
from backend.models import (
    FinalReport,
    SBARData,
    PatientInfo,
    Situation,
    Background,
    Assessment,
    Recommendation,
    Vitals,
    RiskAlert,
)
from backend import database as db
from backend.config import ALLOWED_ORIGINS, REFRESH_TOKEN_EXPIRE_DAYS
from backend.auth import (
    hash_password,
    verify_password,
    create_access_token,
    create_refresh_token,
    decode_token,
    get_current_user,
    get_optional_user,
    require_admin,
    require_nurse_or_admin,
    require_permission,
    require_any_role,
    authenticate_ws_token,
    record_failed_login,
    clear_failed_logins,
    is_account_locked,
    get_lockout_remaining,
    auth_rate_limiter,
    upload_rate_limiter,
)
from backend.constants import ROLES, ROLE_PERMISSIONS, ROLE_DISPLAY, role_has_permission
from backend.middleware import (
    SecurityHeadersMiddleware,
    RateLimitMiddleware,
    RequestLoggingMiddleware,
)
from backend import audio_storage


# ── App lifecycle ─────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    await db.init_db()          # create tables on startup
    yield


app = FastAPI(
    title="MedRelay API",
    version="3.0.0",
    lifespan=lifespan,
    docs_url="/docs",
    redoc_url="/redoc",
)

# ── Middleware (order matters — last added = outermost = first executed) ───────
# Execution order: CORS → Logging → RateLimit → SecurityHeaders → Router
app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(RateLimitMiddleware, max_requests=200, window_seconds=60)
app.add_middleware(RequestLoggingMiddleware)
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-Request-ID"],
    expose_headers=["X-Request-ID"],
)


# ── Health ─────────────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    stats = await db.get_stats()
    return {"status": "ok", "service": "MedRelay", "db_stats": stats}


# ── Demo ───────────────────────────────────────────────────────────────────────

class DemoRequest(BaseModel):
    outgoing: Optional[str] = "Nurse Sarah Chen"
    incoming: Optional[str] = "Nurse Marcus Rivera"


@app.post("/demo")
async def demo_endpoint(request: DemoRequest, _user: dict = Depends(require_permission("create_handoff"))):
    """Run the full 4-agent pipeline on the synthetic demo transcript and persist. Requires authentication."""
    try:
        pipeline = HandoffPipeline()
        final = await pipeline.run_demo(
            outgoing=request.outgoing or "Nurse Sarah Chen",
            incoming=request.incoming or "Nurse Marcus Rivera",
        )
        session_id = await db.save_session(final)
        final.session_id = session_id
        return JSONResponse(content=final.model_dump())
    except Exception as e:
        print(f"[Demo] Error: {traceback.format_exc()}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/demo/patient")
async def demo_patient_template(_user: dict = Depends(get_current_user)):
    """Return a complete demo patient payload with all required SBAR-compatible fields. Requires authentication."""
    return {
        "session_meta": {
            "is_demo": True,
            "outgoing_nurse": "Nurse Sarah Chen, RN",
            "incoming_nurse": "Nurse Marcus Rivera, RN",
            "unit": "ICU",
            "shift": "Night to Day",
            "handoff_type": "Bedside verbal handoff",
        },
        "patient": {
            "name": "Sarah Mitchell",
            "age": "67",
            "mrn": "ICU-2024-0447",
            "room": "ICU 4B",
        },
        "situation": {
            "primary_diagnosis": "Septic shock secondary to pneumonia",
            "reason_for_admission": "Transferred from ED after hypotension, hypoxia, and elevated lactate",
            "current_status": "Hemodynamically unstable, on vasopressor support with high-flow oxygen",
        },
        "background": {
            "relevant_history": "Hypertension, Type 2 diabetes mellitus, CKD stage 2",
            "medications": [
                "Norepinephrine 0.1 mcg/kg/min IV continuous",
                "Vancomycin 1g IV q12h",
                "Piperacillin-Tazobactam 3.375g IV q6h",
                "Heparin 5000 units SubQ q8h (DVT prophylaxis)",
                "Insulin sliding scale AC/HS",
            ],
            "allergies": ["Penicillin (anaphylaxis)", "Latex (rash)"],
            "recent_procedures": [
                "Right subclavian central line placement",
                "Left radial arterial line insertion",
                "Foley catheter insertion",
            ],
        },
        "assessment": {
            "vitals": {
                "bp": "88/54",
                "hr": 118,
                "rr": 24,
                "temp": 38.9,
                "spo2": 91,
            },
            "labs_pending": [
                "Blood cultures x2",
                "Repeat serum lactate (2-hour)",
                "CBC with differential",
                "BMP",
            ],
            "labs_recent": [
                "Lactate 4.2 mmol/L (high)",
                "WBC 18.4 x10^9/L (high)",
                "Procalcitonin 22.1 ng/mL (high)",
                "Creatinine 1.6 mg/dL",
            ],
        },
        "recommendation": {
            "care_plan": "Continue sepsis bundle, titrate norepinephrine to maintain MAP >= 65 mmHg, continue broad-spectrum antibiotics pending cultures",
            "escalation_triggers": "MAP < 65, SpO2 < 88%, altered mental status, urine output < 30 mL/hr",
            "pending_orders": [
                "Repeat lactate in 2 hours",
                "Echocardiogram",
                "Infectious Disease consult",
                "Pharmacy antibiotic review for allergy safety",
            ],
            "next_steps": "Q1h vitals, strict I/O, family update at 0800, reassess pressor need after repeat labs",
        },
        "risk_alert_candidates": [
            {
                "severity": "HIGH",
                "category": "medication",
                "description": "Piperacillin-Tazobactam includes penicillin in a patient with documented penicillin anaphylaxis",
            },
            {
                "severity": "HIGH",
                "category": "vital",
                "description": "SpO2 91% on high-flow oxygen indicates persistent hypoxemia",
            },
            {
                "severity": "HIGH",
                "category": "vital",
                "description": "BP 88/54 suggests ongoing shock despite vasopressor therapy",
            },
        ],
        "transcript_seed": "Patient Sarah Mitchell in ICU 4B, 67-year-old admitted with septic shock from pneumonia. On norepinephrine, low blood pressure 88/54, heart rate 118, O2 sat 91% on high-flow, temp 38.9. Allergic to penicillin. Awaiting blood cultures and repeat lactate. Escalate for MAP below 65 or saturation below 88.",
    }


def _split_pipe(value) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    return [x.strip() for x in text.split("|") if x and x.strip()]


def _as_int(value):
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _as_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _as_str(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _render_imported_report(sbar: SBARData, alerts: list[RiskAlert], outgoing: str, incoming: str, ts: str) -> str:
    vitals = sbar.assessment.vitals
    vitals_line = " | ".join(
        x
        for x in [
            f"BP {vitals.bp}" if vitals.bp else None,
            f"HR {vitals.hr}" if vitals.hr is not None else None,
            f"RR {vitals.rr}" if vitals.rr is not None else None,
            f"Temp {vitals.temp}°C" if vitals.temp is not None else None,
            f"SpO2 {vitals.spo2}%" if vitals.spo2 is not None else None,
        ]
        if x
    ) or "No vitals provided"

    alert_lines = "\n".join(f"[{a.severity}] {a.description}" for a in alerts) if alerts else "No alerts"

    return f"""IMPORTED CLINICAL HANDOFF REPORT
Generated: {ts}
Outgoing: {outgoing} | Incoming: {incoming}

PATIENT: {sbar.patient.name or 'Unknown'} | Age: {sbar.patient.age or 'N/A'} | MRN: {sbar.patient.mrn or 'N/A'} | Room: {sbar.patient.room or 'N/A'}

SITUATION
- Diagnosis: {sbar.situation.primary_diagnosis or 'N/A'}
- Reason for admission: {sbar.situation.reason_for_admission or 'N/A'}
- Current status: {sbar.situation.current_status or 'N/A'}

BACKGROUND
- Relevant history: {sbar.background.relevant_history or 'N/A'}
- Medications: {', '.join(sbar.background.medications) if sbar.background.medications else 'N/A'}
- Allergies: {', '.join(sbar.background.allergies) if sbar.background.allergies else 'N/A'}

ASSESSMENT
- Vitals: {vitals_line}
- Pending labs: {', '.join(sbar.assessment.labs_pending) if sbar.assessment.labs_pending else 'N/A'}
- Recent labs: {', '.join(sbar.assessment.labs_recent) if sbar.assessment.labs_recent else 'N/A'}

RECOMMENDATION
- Care plan: {sbar.recommendation.care_plan or 'N/A'}
- Escalation triggers: {sbar.recommendation.escalation_triggers or 'N/A'}
- Pending orders: {', '.join(sbar.recommendation.pending_orders) if sbar.recommendation.pending_orders else 'N/A'}
- Next steps: {sbar.recommendation.next_steps or 'N/A'}

RISK ALERTS
{alert_lines}
"""


def _build_report_from_feed_row(row: dict, sentinel: SentinelAgent) -> FinalReport:
    outgoing = _as_str(row.get("outgoing_nurse")) or "Nurse Outgoing"
    incoming = _as_str(row.get("incoming_nurse")) or "Nurse Incoming"
    timestamp = _as_str(row.get("timestamp")) or datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    sbar = SBARData(
        patient=PatientInfo(
            name=_as_str(row.get("patient_name")),
            age=_as_str(row.get("patient_age")),
            mrn=_as_str(row.get("patient_mrn")),
            room=_as_str(row.get("patient_room")),
        ),
        situation=Situation(
            primary_diagnosis=_as_str(row.get("primary_diagnosis")),
            reason_for_admission=_as_str(row.get("reason_for_admission")),
            current_status=_as_str(row.get("current_status")),
        ),
        background=Background(
            relevant_history=_as_str(row.get("relevant_history")),
            medications=_split_pipe(row.get("medications")),
            allergies=_split_pipe(row.get("allergies")),
            recent_procedures=_split_pipe(row.get("recent_procedures")),
        ),
        assessment=Assessment(
            vitals=Vitals(
                bp=_as_str(row.get("vitals_bp")),
                hr=_as_int(row.get("vitals_hr")),
                rr=_as_int(row.get("vitals_rr")),
                temp=_as_float(row.get("vitals_temp")),
                spo2=_as_int(row.get("vitals_spo2")),
            ),
            labs_pending=_split_pipe(row.get("labs_pending")),
            labs_recent=_split_pipe(row.get("labs_recent")),
        ),
        recommendation=Recommendation(
            care_plan=_as_str(row.get("care_plan")),
            escalation_triggers=_as_str(row.get("escalation_triggers")),
            pending_orders=_split_pipe(row.get("pending_orders")),
            next_steps=_as_str(row.get("next_steps")),
        ),
    )

    # Sentinel uses deterministic threshold checks and is safe for imported rows.
    # If Sentinel fails for any reason, preserve ingestion with empty alerts.
    alerts: list[RiskAlert] = []
    try:
        # sentinel.check is async; call site will await this via helper endpoint logic
        pass
    except Exception:
        alerts = []

    report = FinalReport(
        sbar=sbar,
        alerts=alerts,
        outgoing_nurse=outgoing,
        incoming_nurse=incoming,
        timestamp=timestamp,
        rendered="",
        is_demo=str(row.get("is_demo", "false")).strip().lower() in {"true", "1", "yes"},
    )
    report.rendered = _render_imported_report(report.sbar, report.alerts, outgoing, incoming, timestamp)
    return report


@app.post("/import/excel")
async def import_excel_feed(file: UploadFile = File(...), dry_run: bool = False, user: dict = Depends(require_permission("import_excel"))):
    """
    Import handoff rows from an Excel file.
    Expected sheet: Patient_Feed_Template (or first sheet fallback).
    """
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        wb = load_workbook(BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    ws = wb["Patient_Feed_Template"] if "Patient_Feed_Template" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="No rows found in sheet")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if not any(headers):
        raise HTTPException(status_code=400, detail="Header row is empty")

    sentinel = SentinelAgent()
    created_ids: list[str] = []
    errors: list[dict] = []

    for idx, values in enumerate(rows[1:], start=2):
        if values is None or not any(v not in (None, "") for v in values):
            continue

        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}

        # Minimal required validation
        if not _as_str(row.get("patient_name")):
            errors.append({"row": idx, "error": "patient_name is required"})
            continue
        if not _as_str(row.get("primary_diagnosis")):
            errors.append({"row": idx, "error": "primary_diagnosis is required"})
            continue

        try:
            report = _build_report_from_feed_row(row, sentinel)
            try:
                report.alerts = await sentinel.check(report.sbar)
            except Exception:
                report.alerts = []
            report.rendered = _render_imported_report(
                report.sbar,
                report.alerts,
                report.outgoing_nurse,
                report.incoming_nurse,
                report.timestamp,
            )

            if not dry_run:
                session_id = await db.save_session(report)
                created_ids.append(session_id)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    return {
        "file": filename,
        "sheet": ws.title,
        "dry_run": dry_run,
        "rows_scanned": max(len(rows) - 1, 0),
        "rows_imported": len(created_ids),
        "session_ids": created_ids,
        "errors": errors,
    }


# ── Sessions REST API ──────────────────────────────────────────────────────────

@app.get("/sessions")
async def list_sessions(limit: int = 50, _user: dict = Depends(require_permission("view_sessions"))):
    """Return summary list of recent handoff sessions. Requires authentication."""
    sessions = await db.get_sessions(limit=limit)
    return {"sessions": sessions, "count": len(sessions)}


@app.get("/sessions/{session_id}")
async def get_session(session_id: str, _user: dict = Depends(require_permission("view_sessions"))):
    """Return full session data including SBAR and rendered report. Requires authentication."""
    session = await db.get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    return session


class SignoffRequest(BaseModel):
    signed_by_outgoing: bool = False
    signed_by_incoming: bool = False


@app.patch("/sessions/{session_id}/signoff")
async def update_signoff(session_id: str, payload: SignoffRequest, _user: dict = Depends(require_permission("sign_off"))):
    """Persist digital sign-off status for a session. Requires authentication."""
    updated = await db.update_signoff(
        session_id,
        outgoing=payload.signed_by_outgoing,
        incoming=payload.signed_by_incoming,
    )
    if not updated:
        raise HTTPException(status_code=404, detail="Session not found")
    return {"session_id": session_id, "updated": True}


@app.delete("/sessions/{session_id}")
async def delete_session(session_id: str, _user: dict = Depends(require_admin)):
    """Delete a session record. Requires admin role."""
    deleted = await db.delete_session(session_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Session not found")
    return {"session_id": session_id, "deleted": True}


@app.get("/stats")
async def stats(_user: dict = Depends(require_permission("view_sessions"))):
    """Return aggregate statistics. Requires view_sessions permission."""
    return await db.get_stats()


@app.get("/analytics")
async def analytics(_user: dict = Depends(require_admin)):
    """Return rich analytics data for the dashboard. Requires admin role."""
    return await db.get_analytics()


@app.get("/patients")
async def list_patients(_user: dict = Depends(require_permission("view_sessions"))):
    """Return unique patients list with session counts. Requires authentication."""
    patients = await db.get_patients()
    return {"patients": patients, "count": len(patients)}


@app.get("/patients/{patient_name}/timeline")
async def patient_timeline(patient_name: str, _user: dict = Depends(require_permission("view_sessions"))):
    """Return all handoff sessions for a specific patient, chronologically. Requires authentication."""
    timeline = await db.get_patient_timeline(patient_name)
    if not timeline:
        raise HTTPException(status_code=404, detail="No sessions found for this patient")
    return {"patient_name": patient_name, "sessions": timeline, "count": len(timeline)}


# ── Authentication API ─────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    username: str = PydanticField(..., min_length=1, max_length=100)
    password: str = PydanticField(..., min_length=1, max_length=200)

    @field_validator("username")
    @classmethod
    def strip_username(cls, v: str) -> str:
        return v.strip()


class RefreshRequest(BaseModel):
    refresh_token: str


class ChangePasswordRequest(BaseModel):
    old_password: str = PydanticField(..., min_length=1)
    new_password: str = PydanticField(..., min_length=6, max_length=200)


@app.post("/auth/login")
async def auth_login(payload: LoginRequest, request: Request):
    """Authenticate user and return JWT access + refresh tokens."""
    ip = request.client.host if request.client else "unknown"

    # Rate limit auth attempts per IP
    if not auth_rate_limiter.is_allowed(ip):
        retry = auth_rate_limiter.get_retry_after(ip)
        raise HTTPException(status_code=429, detail=f"Too many login attempts. Retry in {retry}s.")

    # Check account lockout
    if is_account_locked(payload.username):
        remaining = get_lockout_remaining(payload.username)
        raise HTTPException(
            status_code=423,
            detail=f"Account locked due to too many failed attempts. Try again in {remaining}s.",
        )

    user = await db.admin_login(payload.username, payload.password)
    if not user:
        record_failed_login(payload.username, ip)
        await db.add_audit_log(payload.username, "login_failed", "auth", "", f"IP: {ip}")
        raise HTTPException(status_code=401, detail="Invalid username or password")

    clear_failed_logins(payload.username)

    # Generate tokens
    access_token = create_access_token(user["user_id"], user["username"], user["role"])
    refresh_token = create_refresh_token(user["user_id"])

    # Store refresh token hash
    token_hash = hashlib.sha256(refresh_token.encode()).hexdigest()
    expires_at = (datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)).isoformat()
    await db.store_refresh_token(user["user_id"], token_hash, expires_at)

    await db.add_audit_log(user["username"], "login", "auth", user["user_id"], f"IP: {ip}")

    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "user": {
            "user_id": user["user_id"],
            "username": user["username"],
            "display_name": user.get("display_name", ""),
            "role": user["role"],
            "role_display": ROLE_DISPLAY.get(user["role"], user["role"]),
            "permissions": sorted(list(ROLE_PERMISSIONS.get(user["role"], set()))),
        },
    }


@app.post("/auth/refresh")
async def auth_refresh(payload: RefreshRequest):
    """Exchange a refresh token for a new access token."""
    try:
        token_payload = decode_token(payload.refresh_token)
    except HTTPException:
        raise HTTPException(status_code=401, detail="Invalid or expired refresh token")

    if token_payload.get("type") != "refresh":
        raise HTTPException(status_code=401, detail="Invalid token type")

    user_id = token_payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token subject")

    token_hash = hashlib.sha256(payload.refresh_token.encode()).hexdigest()

    # Validate stored token
    stored = await db.validate_refresh_token(token_hash)
    if not stored:
        raise HTTPException(status_code=401, detail="Refresh token revoked or expired")

    # Get user
    user = await db.get_user_by_id(user_id)
    if not user or not user["is_active"]:
        raise HTTPException(status_code=401, detail="User not found or disabled")

    # Rotate: revoke old, issue new
    await db.revoke_refresh_token(token_hash)

    new_access = create_access_token(user["user_id"], user["username"], user["role"])
    new_refresh = create_refresh_token(user["user_id"])

    new_hash = hashlib.sha256(new_refresh.encode()).hexdigest()
    new_expires = (datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)).isoformat()
    await db.store_refresh_token(user["user_id"], new_hash, new_expires)

    return {
        "access_token": new_access,
        "refresh_token": new_refresh,
        "token_type": "bearer",
        "user": {
            "user_id": user["user_id"],
            "username": user["username"],
            "display_name": user.get("display_name", ""),
            "role": user["role"],
            "role_display": ROLE_DISPLAY.get(user["role"], user["role"]),
            "permissions": sorted(list(ROLE_PERMISSIONS.get(user["role"], set()))),
        },
    }


@app.post("/auth/logout")
async def auth_logout(payload: RefreshRequest, user: dict = Depends(get_current_user)):
    """Revoke the given refresh token."""
    token_hash = hashlib.sha256(payload.refresh_token.encode()).hexdigest()
    await db.revoke_refresh_token(token_hash)
    await db.add_audit_log(user.get("username", ""), "logout", "auth", user.get("sub", ""))
    return {"message": "Logged out successfully"}


@app.post("/auth/logout-all")
async def auth_logout_all(user: dict = Depends(get_current_user)):
    """Revoke all refresh tokens for the current user (sign out everywhere)."""
    count = await db.revoke_all_user_tokens(user["sub"])
    await db.add_audit_log(user.get("username", ""), "logout_all", "auth", user["sub"], f"Revoked {count} tokens")
    return {"message": f"Revoked {count} active sessions"}


@app.post("/auth/change-password")
async def auth_change_password(
    payload: ChangePasswordRequest,
    user: dict = Depends(get_current_user),
):
    """Change the current user's password."""
    ok = await db.change_password(user["sub"], payload.old_password, payload.new_password)
    if not ok:
        raise HTTPException(status_code=400, detail="Current password is incorrect")

    # Revoke all existing refresh tokens (force re-login everywhere)
    await db.revoke_all_user_tokens(user["sub"])
    await db.add_audit_log(user.get("username", ""), "change_password", "auth", user["sub"])

    return {"message": "Password changed successfully. Please log in again."}


@app.get("/auth/me")
async def auth_me(user: dict = Depends(get_current_user)):
    """Return the current authenticated user's profile."""
    full_user = await db.get_user_by_id(user["sub"])
    if not full_user:
        raise HTTPException(status_code=404, detail="User not found")
    role = full_user.get("role", "nurse")
    return {
        "user_id": full_user["user_id"],
        "username": full_user["username"],
        "display_name": full_user.get("display_name", ""),
        "role": role,
        "role_display": ROLE_DISPLAY.get(role, role),
        "permissions": sorted(list(ROLE_PERMISSIONS.get(role, set()))),
        "is_active": full_user.get("is_active", True),
        "created_at": full_user.get("created_at", ""),
        "last_login": full_user.get("last_login", ""),
    }


# ── Admin API ──────────────────────────────────────────────────────────────────

class AdminLoginRequest(BaseModel):
    username: str
    pin: str


@app.post("/admin/login")
async def admin_login(payload: AdminLoginRequest, request: Request):
    """Legacy admin login — redirects to JWT auth. Kept for backward compatibility."""
    ip = request.client.host if request.client else "unknown"

    if not auth_rate_limiter.is_allowed(ip):
        raise HTTPException(status_code=429, detail="Too many login attempts.")

    if is_account_locked(payload.username):
        remaining = get_lockout_remaining(payload.username)
        raise HTTPException(status_code=423, detail=f"Account locked. Try again in {remaining}s.")

    user = await db.admin_login(payload.username, payload.pin)
    if not user:
        record_failed_login(payload.username, ip)
        raise HTTPException(status_code=401, detail="Invalid credentials")

    clear_failed_logins(payload.username)

    access_token = create_access_token(user["user_id"], user["username"], user["role"])
    refresh_token = create_refresh_token(user["user_id"])

    token_hash = hashlib.sha256(refresh_token.encode()).hexdigest()
    expires_at = (datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)).isoformat()
    await db.store_refresh_token(user["user_id"], token_hash, expires_at)

    await db.add_audit_log(payload.username, "login", "admin", user.get("user_id", ""), f"IP: {ip}")
    return {
        "user": {**user, "role_display": ROLE_DISPLAY.get(user.get("role", ""), ""), "permissions": sorted(list(ROLE_PERMISSIONS.get(user.get("role", ""), set())))},
        "authenticated": True,
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
    }


@app.get("/admin/users")
async def list_admin_users(_admin: dict = Depends(require_admin)):
    """Return all admin users. Requires admin role."""
    return {"users": await db.get_admin_users()}


class CreateUserRequest(BaseModel):
    username: str = PydanticField(..., min_length=2, max_length=100)
    display_name: str = ""
    role: str = "nurse"
    password: str = PydanticField(default="changeme", min_length=6, max_length=200)

    @field_validator("role")
    @classmethod
    def validate_role(cls, v: str) -> str:
        if v not in ROLES:
            raise ValueError(f"Invalid role. Must be one of: {', '.join(ROLES)}")
        return v


@app.post("/admin/users")
async def create_admin_user(payload: CreateUserRequest, admin: dict = Depends(require_admin)):
    """Create a new user. Requires admin role."""
    try:
        user = await db.create_admin_user(payload.username, payload.display_name, payload.role, payload.password)
        await db.add_audit_log(admin.get("username", "admin"), "create_user", "user", user["user_id"], f"Created user {payload.username}")
        return user
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class UpdateUserRequest(BaseModel):
    display_name: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None


@app.patch("/admin/users/{user_id}")
async def update_admin_user(user_id: str, payload: UpdateUserRequest, admin: dict = Depends(require_admin)):
    """Update a user's profile. Requires admin role."""
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    ok = await db.update_admin_user(user_id, **updates)
    if not ok:
        raise HTTPException(status_code=404, detail="User not found")
    await db.add_audit_log(admin.get("username", "admin"), "update_user", "user", user_id, str(updates))
    return {"user_id": user_id, "updated": True}


@app.delete("/admin/users/{user_id}")
async def delete_admin_user(user_id: str, admin: dict = Depends(require_admin)):
    """Delete a user. Requires admin role."""
    ok = await db.delete_admin_user(user_id)
    if not ok:
        raise HTTPException(status_code=404, detail="User not found")
    # Revoke all their tokens
    await db.revoke_all_user_tokens(user_id)
    await db.add_audit_log(admin.get("username", "admin"), "delete_user", "user", user_id)
    return {"user_id": user_id, "deleted": True}


@app.get("/admin/audit")
async def get_audit_log(limit: int = 100, _user: dict = Depends(require_permission("view_audit"))):
    """Return recent audit log entries. Requires view_audit permission."""
    logs = await db.get_audit_log(limit=limit)
    return {"logs": logs, "count": len(logs)}


@app.get("/admin/settings")
async def get_settings(_admin: dict = Depends(require_admin)):
    """Return all system settings. Requires admin role."""
    return await db.get_settings()


class UpdateSettingsRequest(BaseModel):
    settings: dict


@app.put("/admin/settings")
async def update_settings(payload: UpdateSettingsRequest, admin: dict = Depends(require_admin)):
    """Update system settings. Requires admin role."""
    await db.update_settings(payload.settings, updated_by=admin.get("username", "admin"))
    await db.add_audit_log(admin.get("username", "admin"), "update_settings", "system", "", str(payload.settings))
    return {"updated": True}


class BulkDeleteRequest(BaseModel):
    session_ids: list[str]


@app.post("/admin/sessions/bulk-delete")
async def bulk_delete_sessions(payload: BulkDeleteRequest, user: dict = Depends(require_permission("manage_sessions"))):
    """Delete multiple sessions at once. Requires manage_sessions permission."""
    count = await db.bulk_delete_sessions(payload.session_ids)
    await db.add_audit_log(user.get("username", "admin"), "bulk_delete", "sessions", "", f"Deleted {count} sessions")
    return {"deleted": count}


@app.post("/admin/sessions/purge-demos")
async def purge_demo_sessions(dry_run: bool = True, user: dict = Depends(require_permission("manage_sessions"))):
    """Delete all demo sessions. Defaults to dry-run for safety. Requires manage_sessions permission."""
    stats = await db.get_stats()
    would_delete = int(stats.get("demo_sessions", 0) or 0)

    if dry_run:
        return {
            "dry_run": True,
            "would_delete": would_delete,
            "deleted": 0,
        }

    count = await db.purge_demo_sessions()
    await db.add_audit_log(user.get("username", "admin"), "purge_demos", "sessions", "", f"Purged {count} demo sessions")
    return {"dry_run": False, "deleted": count}


# ── Roles & Permissions API ───────────────────────────────────────────────────

@app.get("/roles")
async def list_roles(_user: dict = Depends(get_current_user)):
    """Return available roles and their permissions."""
    return {
        "roles": [
            {
                "key": role,
                "display": ROLE_DISPLAY.get(role, role),
                "permissions": sorted(list(ROLE_PERMISSIONS.get(role, set()))),
            }
            for role in ROLES
        ],
    }


@app.get("/auth/permissions")
async def my_permissions(user: dict = Depends(get_current_user)):
    """Return the current user's role and associated permissions."""
    role = user.get("role", "nurse")
    return {
        "role": role,
        "role_display": ROLE_DISPLAY.get(role, role),
        "permissions": sorted(list(ROLE_PERMISSIONS.get(role, set()))),
    }


# ── Enhanced Analytics Endpoints ──────────────────────────────────────────────

@app.get("/analytics/nurses")
async def analytics_nurse_performance(_user: dict = Depends(require_admin)):
    """Return per-nurse handoff performance metrics. Requires admin role."""
    return await db.get_nurse_analytics()


@app.get("/analytics/trends")
async def analytics_trends(_user: dict = Depends(require_admin)):
    """Return weekly/monthly trend data with comparisons. Requires admin role."""
    return await db.get_trend_analytics()


@app.get("/analytics/quality")
async def analytics_quality(_user: dict = Depends(require_admin)):
    """Return handoff quality scores based on completeness and sign-off compliance. Requires admin role."""
    return await db.get_quality_analytics()


# ═══════════════════════════════════════════════════════════════════════════════
#  NURSE SCHEDULING API
# ═══════════════════════════════════════════════════════════════════════════════

from backend.models import (
    PatientRegistryCreate, PatientRegistryUpdate,
    ScheduleCreate, ScheduleUpdate,
    AssignmentCreate, AutoScheduleRequest,
)

# ── Scheduling Stats ──────────────────────────────────────────────────────────

@app.get("/scheduling/stats")
async def scheduling_stats(_user: dict = Depends(get_current_user)):
    """Return an overview of scheduling: admitted patients, active nurses, schedule counts."""
    return await db.get_schedule_stats()


# ── Patient Registry CRUD ─────────────────────────────────────────────────────

@app.get("/scheduling/patients")
async def list_registry_patients(status: str | None = None, _user: dict = Depends(get_current_user)):
    """List patients in the hospital registry. Optional ?status=admitted|discharged|transferred."""
    patients = await db.get_patients_registry(status=status)
    return {"patients": patients, "count": len(patients)}


@app.post("/scheduling/patients")
async def create_registry_patient(body: PatientRegistryCreate, _user: dict = Depends(require_admin)):
    """Register a new patient (admin only)."""
    result = await db.create_patient(
        name=body.name, mrn=body.mrn, room=body.room, bed=body.bed,
        acuity=body.acuity, diagnosis=body.diagnosis, notes=body.notes,
        admission_date=body.admission_date,
    )
    await db.add_audit_log(_user["username"], "patient_created", "patient", result["patient_id"],
                           f"Registered patient {body.name}")
    return result


@app.get("/scheduling/patients/{patient_id}")
async def get_registry_patient(patient_id: str, _user: dict = Depends(get_current_user)):
    """Get a single patient from the registry."""
    patient = await db.get_patient_by_id(patient_id)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    return patient


@app.put("/scheduling/patients/{patient_id}")
async def update_registry_patient(patient_id: str, body: PatientRegistryUpdate, _user: dict = Depends(require_admin)):
    """Update patient info (admin only)."""
    updates = body.model_dump(exclude_none=True)
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    ok = await db.update_patient(patient_id, **updates)
    if not ok:
        raise HTTPException(status_code=404, detail="Patient not found")
    await db.add_audit_log(_user["username"], "patient_updated", "patient", patient_id,
                           f"Updated: {', '.join(updates.keys())}")
    return {"ok": True}


@app.delete("/scheduling/patients/{patient_id}")
async def delete_registry_patient(patient_id: str, _user: dict = Depends(require_admin)):
    """Remove a patient from the registry (admin only)."""
    ok = await db.delete_patient(patient_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Patient not found")
    await db.add_audit_log(_user["username"], "patient_deleted", "patient", patient_id, "")
    return {"ok": True}


# ── Schedules CRUD ────────────────────────────────────────────────────────────

@app.get("/scheduling/schedules")
async def list_schedules(shift_date: str | None = None, status: str | None = None,
                         _user: dict = Depends(get_current_user)):
    """List schedules with optional date/status filters."""
    schedules = await db.get_schedules(shift_date=shift_date, status=status)
    return {"schedules": schedules, "count": len(schedules)}


@app.post("/scheduling/schedules")
async def create_schedule_endpoint(body: ScheduleCreate, user: dict = Depends(require_admin)):
    """Create a new schedule draft (admin only)."""
    result = await db.create_schedule(
        shift_date=body.shift_date, shift_type=body.shift_type,
        created_by=user["sub"], notes=body.notes,
    )
    await db.add_audit_log(user["username"], "schedule_created", "schedule", result["schedule_id"],
                           f"{body.shift_date} {body.shift_type}")
    return result


@app.get("/scheduling/schedules/{schedule_id}")
async def get_schedule_endpoint(schedule_id: str, _user: dict = Depends(get_current_user)):
    """Get a schedule with all its nurse–patient assignments."""
    sched = await db.get_schedule(schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schedule not found")
    return sched


@app.put("/scheduling/schedules/{schedule_id}")
async def update_schedule_endpoint(schedule_id: str, body: ScheduleUpdate, user: dict = Depends(require_admin)):
    """Update schedule status/notes (admin only). Use status='published' to finalize."""
    updates = body.model_dump(exclude_none=True)
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    if updates.get("status") == "published":
        updates["published_at"] = datetime.now().isoformat()
    ok = await db.update_schedule(schedule_id, **updates)
    if not ok:
        raise HTTPException(status_code=404, detail="Schedule not found")
    await db.add_audit_log(user["username"], "schedule_updated", "schedule", schedule_id,
                           f"Updated: {', '.join(updates.keys())}")
    return {"ok": True}


@app.delete("/scheduling/schedules/{schedule_id}")
async def delete_schedule_endpoint(schedule_id: str, user: dict = Depends(require_admin)):
    """Delete a schedule and all its assignments (admin only)."""
    ok = await db.delete_schedule(schedule_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Schedule not found")
    await db.add_audit_log(user["username"], "schedule_deleted", "schedule", schedule_id, "")
    return {"ok": True}


# ── Assignments ───────────────────────────────────────────────────────────────

@app.post("/scheduling/schedules/{schedule_id}/assignments")
async def add_assignment_endpoint(schedule_id: str, body: AssignmentCreate,
                                  user: dict = Depends(require_admin)):
    """Manually assign a nurse to a patient within a schedule (admin only)."""
    result = await db.add_assignment(
        schedule_id=schedule_id, nurse_user_id=body.nurse_user_id,
        patient_id=body.patient_id, is_primary=body.is_primary, notes=body.notes,
    )
    await db.add_audit_log(user["username"], "assignment_created", "assignment", result["assignment_id"],
                           f"nurse={body.nurse_user_id}, patient={body.patient_id}")
    return result


@app.delete("/scheduling/assignments/{assignment_id}")
async def remove_assignment_endpoint(assignment_id: str, user: dict = Depends(require_admin)):
    """Remove a single nurse–patient assignment (admin only)."""
    ok = await db.remove_assignment(assignment_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Assignment not found")
    await db.add_audit_log(user["username"], "assignment_deleted", "assignment", assignment_id, "")
    return {"ok": True}



@app.get("/scheduling/ai-analysis")
async def schedule_analysis(schedule_id: str | None = None, user: dict = Depends(require_nurse_or_admin)):
    """
    Agent 10: Staffing Agent
    Analyze the current schedule and patient acuity to provide staffing recommendations.
    """
    # 1. Get Patients (admitted)
    patients = await db.get_patients_registry(status="admitted")

    # 2. Get Nurses (active)
    users = await db.get_admin_users()
    # Filter only relevant roles and respect is_active flag
    # shift_status is now considered by the agent (active vs absent)
    nurses = [u for u in users if u["role"] in ("nurse", "charge_nurse", "supervisor") and u["is_active"]]

    # 3. Get assignments
    assignments = []
    if schedule_id:
        schedule = await db.get_schedule(schedule_id)
        if schedule:
            assignments = schedule.get("assignments", [])

    # 4. Compute risk scores
    risk_data = {}
    for p in patients:
        # Check history for trends
        hist = await db.get_history_for_trends(p.get("mrn", ""), p.get("name", ""), limit=1)
        score = 0
        
        if hist and len(hist) > 0:
            last = hist[0]
            h = last.get("high_alert_count", 0) or 0
            m = last.get("medium_alert_count", 0) or 0
            l = last.get("low_alert_count", 0) or 0
            score = min(100, (h * 20) + (m * 8) + (l * 2))
            # Boost if no alerts but high acuity manual
            if score < 20 and p.get("acuity", 1) >= 4:
                score = 75
        else:
            # Fallback to manual acuity
            acuity = p.get("acuity", 3)
            score = acuity * 18  # 1=18, 5=90

        risk_data[p["patient_id"]] = {"score": score, "alerts": []}

    try:
        agent = StaffingAgent()
        analysis = await agent.analyze(nurses, patients, assignments, risk_data)
        return analysis
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))



@app.get("/analytics/briefing")
async def get_analytics_briefing(user: dict = Depends(require_nurse_or_admin)):
    """
    Agent 12: CMIO Agent
    Generates a daily executive summary of unit performance, risk, and revenue.
    """
    # 1. Gather all data
    stats = await db.get_analytics()
    recent_alerts = await db.get_recent_critical_alerts(limit=8)
    # Get revenue estimate from recent sessions (naive sum from billing_json if available)
    # For now, we'll let the agent infer from the high_alert_count or simulated data if billing isn't populated yet.
    
    # 2. Call Agent
    try:
        agent = CMIOAgent()
        briefing = await agent.generate_briefing(stats, recent_alerts)
        return briefing
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ── Auto-Scheduling ──────────────────────────────────────────────────────────

@app.post("/scheduling/schedules/{schedule_id}/auto")
async def auto_schedule_endpoint(schedule_id: str, body: AutoScheduleRequest,
                                 user: dict = Depends(require_admin)):
    """Run the acuity-balanced auto-scheduler for a schedule (admin only).
    Replaces all existing assignments with optimally balanced ones."""
    # Verify schedule exists
    sched = await db.get_schedule(schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schedule not found")
    if sched["status"] == "archived":
        raise HTTPException(status_code=400, detail="Cannot modify an archived schedule")

    result = await db.auto_schedule(schedule_id, max_patients_per_nurse=body.max_patients_per_nurse)
    if "error" in result and result.get("assigned", 0) == 0:
        raise HTTPException(status_code=400, detail=result["error"])
    await db.add_audit_log(user["username"], "auto_schedule", "schedule", schedule_id,
                           f"Assigned {result['assigned']}/{result['total_patients']} patients to {result['total_nurses']} nurses")
    return result


# ── My Schedule (for nurses) ─────────────────────────────────────────────────

@app.get("/scheduling/my-schedule")
async def my_schedule(shift_date: str | None = None, user: dict = Depends(get_current_user)):
    """Get the current user's patient assignments. Nurses see their own schedule."""
    assignments = await db.get_nurse_schedule(user["sub"], shift_date=shift_date)
    return {"assignments": assignments, "count": len(assignments)}


@app.put("/scheduling/staff/{user_id}/status")
async def update_staff_status(user_id: str, body: dict, user: dict = Depends(require_admin)):
    """Update a nurse's shift status (active, absent, break, on_call)."""
    status = body.get("status", "active")
    if status not in ("active", "absent", "break", "on_call"):
        raise HTTPException(status_code=400, detail="Invalid status")
    
    ok = await db.update_admin_user(user_id, shift_status=status)
    if not ok:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.add_audit_log(user["username"], "update_staff_status", "user", user_id, f"Set status to {status}")
    return {"ok": True, "user_id": user_id, "status": status}


@app.put("/scheduling/assignments/{assignment_id}/handoff-complete")
async def mark_handoff_complete(assignment_id: str, user: dict = Depends(get_current_user)):
    """Mark a schedule assignment's handoff as completed."""
    ok = await db.mark_assignment_handoff_complete(assignment_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Assignment not found")
    await db.add_audit_log(user["username"], "handoff_completed", "assignment", assignment_id, "")
    return {"ok": True, "assignment_id": assignment_id, "handoff_status": "completed"}


@app.get("/scheduling/patients/{patient_id}/previous-nurse")
async def previous_nurse_endpoint(patient_id: str, shift_date: str, shift_type: str,
                                  _user: dict = Depends(get_current_user)):
    """Find which nurse had this patient in the previous shift (for handoff auto-fill)."""
    result = await db.get_previous_shift_nurse(patient_id, shift_date, shift_type)
    if result:
        return {"found": True, **result}
    return {"found": False, "nurse_user_id": None, "nurse_name": None, "shift_date": shift_date, "shift_type": shift_type}


# ── Recordings API (Local Audio Storage) ──────────────────────────────────────

@app.get("/recordings")
async def list_recordings(limit: int = 50, _user: dict = Depends(require_permission("view_sessions"))):
    """List locally stored audio recordings with metadata."""
    recordings = await audio_storage.list_recordings(limit=limit)
    return {"recordings": recordings, "count": len(recordings)}


@app.get("/recordings/{recording_id}")
async def get_recording_info(recording_id: str, _user: dict = Depends(require_permission("view_sessions"))):
    """Return metadata for a specific recording."""
    meta = await audio_storage.get_recording_metadata(recording_id)
    if not meta:
        raise HTTPException(status_code=404, detail="Recording not found")
    transcript = await audio_storage.get_transcript(recording_id)
    meta["transcript"] = transcript
    return meta


@app.get("/recordings/{recording_id}/audio")
async def download_recording(recording_id: str, _user: dict = Depends(require_permission("view_sessions"))):
    """Download the WAV audio file for playback or re-processing."""
    wav_data = await audio_storage.get_recording_wav(recording_id)
    if not wav_data:
        raise HTTPException(status_code=404, detail="Recording WAV file not found")
    return Response(
        content=wav_data,
        media_type="audio/wav",
        headers={"Content-Disposition": f'attachment; filename="{recording_id}.wav"'},
    )


@app.get("/recordings/{recording_id}/transcript")
async def get_recording_transcript(recording_id: str, _user: dict = Depends(require_permission("view_sessions"))):
    """Return the saved transcript for a recording."""
    transcript = await audio_storage.get_transcript(recording_id)
    if transcript is None:
        raise HTTPException(status_code=404, detail="Transcript not found for this recording")
    return {"recording_id": recording_id, "transcript": transcript}


@app.post("/recordings/{recording_id}/reprocess")
async def reprocess_recording(
    recording_id: str,
    outgoing: str = "",
    incoming: str = "",
    _user: dict = Depends(require_permission("create_handoff"))
):
    """
    Re-process a saved recording through the full pipeline.
    Loads the saved WAV, re-transcribes, and runs Extract → Sentinel → Bridge.
    Useful when AI models improve or when initial transcription was poor.
    """
    wav_data = await audio_storage.get_recording_wav(recording_id)
    if not wav_data:
        raise HTTPException(status_code=404, detail="Recording WAV file not found")

    meta = await audio_storage.get_recording_metadata(recording_id) or {}
    out_nurse = outgoing or meta.get("outgoing_nurse", "Nurse Outgoing")
    in_nurse = incoming or meta.get("incoming_nurse", "Nurse Incoming")

    try:
        # Re-transcribe from saved WAV
        relay = RelayAgent()
        await relay.process_audio_chunk(wav_data)
        transcript = await relay.transcribe_full()

        if not transcript or not transcript.strip():
            return JSONResponse(status_code=422, content={
                "error": "Re-transcription produced no text",
                "recording_id": recording_id,
            })

        # Save updated transcript
        await audio_storage.save_transcript(recording_id, transcript)

        # Run full pipeline
        pipeline = HandoffPipeline()
        final = await pipeline.run_from_transcript(transcript, out_nurse, in_nurse)
        session_id = await db.save_session(final)
        final.session_id = session_id

        return JSONResponse(content={
            "recording_id": recording_id,
            "session_id": session_id,
            "transcript_length": len(transcript),
            "report": final.model_dump(),
        })
    except Exception as e:
        print(f"[Reprocess] Error: {traceback.format_exc()}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.delete("/recordings/{recording_id}")
async def delete_recording(recording_id: str, _user: dict = Depends(require_admin)):
    """Delete a recording and all associated files. Requires admin role."""
    deleted = await audio_storage.delete_recording(recording_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Recording not found")
    return {"recording_id": recording_id, "deleted": True}


# ── Shift Summary API ─────────────────────────────────────────────────────────

@app.get("/shift-summary")
async def shift_summary(
    nurse: Optional[str] = None,
    date: Optional[str] = None,
    shift: Optional[str] = None,
    _user: dict = Depends(require_permission("view_sessions")),
):
    """
    Aggregate multiple handoffs into a consolidated shift summary.
    Filters by nurse, date, and shift period (day/night/all).
    Returns a structured overview of all patients handed off during the shift.
    """
    try:
        # Determine date range
        target_date = date or datetime.now().strftime("%Y-%m-%d")
        shift = (shift or "all").lower()

        sessions = await db.get_sessions(limit=200)

        # Filter sessions by date and optionally by nurse
        filtered = []
        for s in sessions:
            ts = s.get("timestamp", "") or s.get("created_at", "")
            if not ts.startswith(target_date):
                continue

            # Filter by shift period
            if shift != "all":
                try:
                    hour = int(ts[11:13]) if len(ts) > 13 else 12
                    if shift == "day" and (hour < 7 or hour >= 19):
                        continue
                    if shift == "night" and (7 <= hour < 19):
                        continue
                except (ValueError, IndexError):
                    pass

            # Filter by nurse name (either outgoing or incoming)
            if nurse:
                nurse_lower = nurse.lower()
                out_nurse = (s.get("outgoing_nurse") or "").lower()
                in_nurse = (s.get("incoming_nurse") or "").lower()
                if nurse_lower not in out_nurse and nurse_lower not in in_nurse:
                    continue

            filtered.append(s)

        # Build summary
        patients = {}
        total_high_alerts = 0
        total_medium_alerts = 0
        total_low_alerts = 0
        nurses_involved = set()

        for s in filtered:
            p_name = s.get("patient_name", "Unknown")
            if p_name not in patients:
                patients[p_name] = {
                    "name": p_name,
                    "room": s.get("patient_room", ""),
                    "mrn": s.get("patient_mrn", ""),
                    "diagnosis": s.get("diagnosis", ""),
                    "handoff_count": 0,
                    "latest_timestamp": "",
                    "high_alerts": 0,
                    "signed_off": False,
                }
            patients[p_name]["handoff_count"] += 1
            patients[p_name]["latest_timestamp"] = s.get("timestamp", "")
            patients[p_name]["high_alerts"] += s.get("high_alert_count", 0)
            patients[p_name]["signed_off"] = bool(s.get("signed_by_outgoing") and s.get("signed_by_incoming"))

            total_high_alerts += s.get("high_alert_count", 0)
            total_medium_alerts += s.get("medium_alert_count", 0)
            total_low_alerts += s.get("low_alert_count", 0)
            nurses_involved.add(s.get("outgoing_nurse", ""))
            nurses_involved.add(s.get("incoming_nurse", ""))

        nurses_involved.discard("")

        # Sort patients: HIGH alert patients first
        patient_list = sorted(patients.values(), key=lambda p: -p["high_alerts"])

        return {
            "date": target_date,
            "shift": shift,
            "nurse_filter": nurse,
            "total_handoffs": len(filtered),
            "total_patients": len(patients),
            "total_high_alerts": total_high_alerts,
            "total_medium_alerts": total_medium_alerts,
            "total_low_alerts": total_low_alerts,
            "nurses_involved": sorted(list(nurses_involved)),
            "patients": patient_list,
            "sign_off_rate": round(
                sum(1 for p in patient_list if p["signed_off"]) / max(len(patient_list), 1) * 100
            ),
        }
    except Exception as e:
        print(f"[ShiftSummary] Error: {traceback.format_exc()}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ── PDF Export API ─────────────────────────────────────────────────────────────

@app.get("/sessions/{session_id}/pdf")
async def export_session_pdf(session_id: str, _user: dict = Depends(require_permission("view_sessions"))):
    """
    Export a session's SBAR report as a downloadable PDF.
    Uses reportlab for PDF generation (lightweight, no external deps).
    """
    session = await db.get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    try:
        pdf_bytes = _generate_pdf(session)
        patient_name = (session.get("patient_name") or "report").replace(" ", "_")
        filename = f"MedRelay_{patient_name}_{session_id[:8]}.pdf"

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ImportError:
        # Fallback: return plain text if reportlab is not installed
        rendered = session.get("rendered", "No report available")
        return Response(
            content=rendered.encode("utf-8"),
            media_type="text/plain",
            headers={"Content-Disposition": f'attachment; filename="MedRelay_{session_id[:8]}.txt"'},
        )
    except Exception as e:
        print(f"[PDF] Error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")


def _generate_pdf(session: dict) -> bytes:
    """Generate a PDF from session data using reportlab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm,
                            leftMargin=20*mm, rightMargin=20*mm)
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle("CustomTitle", parent=styles["Title"], fontSize=18,
                                  textColor=HexColor("#1e293b"), spaceAfter=6)
    heading_style = ParagraphStyle("CustomHeading", parent=styles["Heading2"], fontSize=13,
                                    textColor=HexColor("#4338ca"), spaceBefore=12, spaceAfter=4)
    body_style = ParagraphStyle("CustomBody", parent=styles["Normal"], fontSize=10,
                                 leading=14, textColor=HexColor("#334155"))
    alert_high = ParagraphStyle("AlertHigh", parent=body_style, textColor=HexColor("#dc2626"),
                                 fontName="Helvetica-Bold")
    alert_medium = ParagraphStyle("AlertMedium", parent=body_style, textColor=HexColor("#d97706"))
    alert_low = ParagraphStyle("AlertLow", parent=body_style, textColor=HexColor("#2563eb"))
    meta_style = ParagraphStyle("Meta", parent=body_style, fontSize=9, textColor=HexColor("#64748b"))

    elements = []

    # Header
    elements.append(Paragraph("MedRelay Clinical Handoff Report", title_style))
    elements.append(Paragraph(
        f"Session: {session.get('session_id', 'N/A')[:12]}… | "
        f"Date: {session.get('timestamp', 'N/A')} | "
        f"{'DEMO' if session.get('is_demo') else 'LIVE'}",
        meta_style
    ))
    elements.append(Spacer(1, 8))

    # Nurses
    elements.append(Paragraph(
        f"Outgoing: {session.get('outgoing_nurse', 'N/A')} → "
        f"Incoming: {session.get('incoming_nurse', 'N/A')}",
        body_style
    ))
    elements.append(Spacer(1, 6))

    # SBAR sections
    sbar = session.get("sbar_json") or session.get("sbar") or {}
    if isinstance(sbar, str):
        import json as _json
        try:
            sbar = _json.loads(sbar)
        except Exception:
            sbar = {}

    # Patient
    patient = sbar.get("patient", {})
    elements.append(Paragraph("PATIENT", heading_style))
    elements.append(Paragraph(
        f"Name: {patient.get('name', 'N/A')} | Age: {patient.get('age', 'N/A')} | "
        f"MRN: {patient.get('mrn', 'N/A')} | Room: {patient.get('room', 'N/A')}",
        body_style
    ))

    # Situation
    sit = sbar.get("situation", {})
    elements.append(Paragraph("SITUATION", heading_style))
    elements.append(Paragraph(f"Diagnosis: {sit.get('primary_diagnosis', 'N/A')}", body_style))
    elements.append(Paragraph(f"Reason: {sit.get('reason_for_admission', 'N/A')}", body_style))
    elements.append(Paragraph(f"Status: {sit.get('current_status', 'N/A')}", body_style))

    # Background
    bg = sbar.get("background", {})
    elements.append(Paragraph("BACKGROUND", heading_style))
    elements.append(Paragraph(f"History: {bg.get('relevant_history', 'N/A')}", body_style))
    meds = bg.get("medications", [])
    elements.append(Paragraph(f"Medications: {', '.join(meds) if meds else 'N/A'}", body_style))
    allergies = bg.get("allergies", [])
    allergy_text = ', '.join(allergies) if allergies else 'N/A'
    if allergies:
        elements.append(Paragraph(f"⚠ Allergies: {allergy_text}", alert_high))
    else:
        elements.append(Paragraph(f"Allergies: {allergy_text}", body_style))

    # Assessment
    assess = sbar.get("assessment", {})
    vitals = assess.get("vitals", {})
    elements.append(Paragraph("ASSESSMENT", heading_style))
    vitals_str = " | ".join(filter(None, [
        f"BP {vitals.get('bp')}" if vitals.get("bp") else None,
        f"HR {vitals.get('hr')}" if vitals.get("hr") else None,
        f"RR {vitals.get('rr')}" if vitals.get("rr") else None,
        f"Temp {vitals.get('temp')}°C" if vitals.get("temp") else None,
        f"SpO2 {vitals.get('spo2')}%" if vitals.get("spo2") else None,
    ])) or "No vitals recorded"
    elements.append(Paragraph(f"Vitals: {vitals_str}", body_style))

    # Recommendation
    rec = sbar.get("recommendation", {})
    elements.append(Paragraph("RECOMMENDATION", heading_style))
    elements.append(Paragraph(f"Care Plan: {rec.get('care_plan', 'N/A')}", body_style))
    elements.append(Paragraph(f"Escalation: {rec.get('escalation_triggers', 'N/A')}", body_style))
    pending = rec.get("pending_orders", [])
    elements.append(Paragraph(f"Pending: {', '.join(pending) if pending else 'N/A'}", body_style))

    # Risk Alerts
    alerts = session.get("alerts_json") or session.get("alerts") or []
    if isinstance(alerts, str):
        import json as _json
        try:
            alerts = _json.loads(alerts)
        except Exception:
            alerts = []

    if alerts:
        elements.append(Paragraph("RISK ALERTS", heading_style))
        for alert in alerts:
            severity = alert.get("severity", "LOW")
            desc = alert.get("description", "")
            style = alert_high if severity == "HIGH" else alert_medium if severity == "MEDIUM" else alert_low
            elements.append(Paragraph(f"[{severity}] {desc}", style))

    # Sign-off status
    elements.append(Spacer(1, 12))
    signed_out = "✓" if session.get("signed_by_outgoing") else "✗"
    signed_in = "✓" if session.get("signed_by_incoming") else "✗"
    elements.append(Paragraph(
        f"Sign-off: Outgoing {signed_out} | Incoming {signed_in}",
        meta_style
    ))

    # Footer
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        "Generated by MedRelay Clinical Intelligence Platform",
        ParagraphStyle("Footer", parent=meta_style, alignment=TA_CENTER, fontSize=8)
    ))

    doc.build(elements)
    return buffer.getvalue()


@app.get("/shift-summary/pdf")
async def export_shift_summary_pdf(
    nurse: Optional[str] = None,
    date: Optional[str] = None,
    shift: Optional[str] = None,
    _user: dict = Depends(require_permission("view_sessions")),
):
    """Export the shift summary as a PDF."""
    # Reuse the shift-summary logic
    from starlette.testclient import TestClient
    summary_data = await shift_summary(nurse=nurse, date=date, shift=shift, _user=_user)

    if isinstance(summary_data, JSONResponse):
        raise HTTPException(status_code=500, detail="Failed to generate shift summary")

    try:
        pdf_bytes = _generate_shift_pdf(summary_data)
        target_date = summary_data.get("date", "report")
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="ShiftSummary_{target_date}.pdf"'},
        )
    except ImportError:
        return JSONResponse(content=summary_data)
    except Exception as e:
        print(f"[PDF] Shift summary PDF error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


def _generate_shift_pdf(summary: dict) -> bytes:
    """Generate a PDF for the shift summary."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm,
                            leftMargin=20*mm, rightMargin=20*mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=16,
                                  textColor=HexColor("#1e293b"))
    heading_style = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=12,
                                    textColor=HexColor("#4338ca"), spaceBefore=10)
    body_style = ParagraphStyle("Body2", parent=styles["Normal"], fontSize=10,
                                 textColor=HexColor("#334155"))
    alert_style = ParagraphStyle("Alert2", parent=body_style, textColor=HexColor("#dc2626"),
                                  fontName="Helvetica-Bold")
    meta_style = ParagraphStyle("Meta2", parent=body_style, fontSize=9, textColor=HexColor("#64748b"))

    elements = []
    elements.append(Paragraph("MedRelay Shift Summary", title_style))
    elements.append(Paragraph(
        f"Date: {summary.get('date')} | Shift: {summary.get('shift', 'all').title()} | "
        f"Nurse: {summary.get('nurse_filter') or 'All'}",
        meta_style
    ))
    elements.append(Spacer(1, 8))

    # Stats
    elements.append(Paragraph("OVERVIEW", heading_style))
    elements.append(Paragraph(
        f"Total Handoffs: {summary.get('total_handoffs', 0)} | "
        f"Patients: {summary.get('total_patients', 0)} | "
        f"Sign-off Rate: {summary.get('sign_off_rate', 0)}%",
        body_style
    ))
    if summary.get("total_high_alerts", 0):
        elements.append(Paragraph(
            f"⚠ HIGH Alerts: {summary['total_high_alerts']}",
            alert_style
        ))

    # Patient list
    patients = summary.get("patients", [])
    if patients:
        elements.append(Paragraph("PATIENTS", heading_style))
        for p in patients:
            alert_marker = f" ⚠ {p['high_alerts']} HIGH" if p.get("high_alerts") else ""
            signed = "✓" if p.get("signed_off") else "✗"
            elements.append(Paragraph(
                f"• {p['name']} | Room: {p.get('room', 'N/A')} | "
                f"Dx: {p.get('diagnosis', 'N/A')} | "
                f"Handoffs: {p.get('handoff_count', 0)} | Signed: {signed}{alert_marker}",
                body_style
            ))

    # Nurses
    nurses = summary.get("nurses_involved", [])
    if nurses:
        elements.append(Paragraph("NURSES INVOLVED", heading_style))
        elements.append(Paragraph(", ".join(nurses), body_style))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        "Generated by MedRelay Clinical Intelligence Platform",
        ParagraphStyle("Footer2", parent=meta_style, alignment=TA_CENTER, fontSize=8)
    ))

    doc.build(elements)
    return buffer.getvalue()


# ── WebSocket ─────────────────────────────────────────────────────────────────

# Live partial transcription every N audio chunks (~2s each)
_LIVE_TRANSCRIBE_EVERY = 8  # 8 chunks ≈ 16 seconds — balances responsiveness vs CPU load

@app.websocket("/ws/handoff")
async def websocket_endpoint(websocket: WebSocket, token: Optional[str] = Query(None)):
    """
    WebSocket flow (optionally authenticated via ?token=<jwt>):
      1. {"type":"start","outgoing":"...","incoming":"..."} - initialise
      2. Binary audio chunks                               - stream audio
         (partial transcripts streamed back every ~10s)
      3. {"type":"end"}                                    - full transcribe → pipeline
    OR
      {"type":"demo","outgoing":"...","incoming":"..."}    - skip to demo
    """
    # Optional authentication — if a token is provided, verify it
    ws_user = None
    if token:
        try:
            ws_user = authenticate_ws_token(token)
        except ValueError:
            pass  # stale / invalid token — proceed unauthenticated

    await websocket.accept()
    audio_chunks: list[bytes] = []
    outgoing_nurse = ""
    incoming_nurse = ""
    language = "en"
    pipeline = HandoffPipeline()
    last_transcribed_idx = 0

    try:
        while True:
            message = await websocket.receive()

            if "text" in message:
                try:
                    data = json.loads(message["text"])
                except json.JSONDecodeError:
                    continue

                msg_type = data.get("type")

                if msg_type == "start":
                    outgoing_nurse = data.get("outgoing", "")
                    incoming_nurse = data.get("incoming", "")
                    language = data.get("language", "en")
                    audio_chunks = []
                    last_transcribed_idx = 0
                    await websocket.send_json({
                        "stage": "listening",
                        "message": "Listening started — speak clearly near microphone.",
                    })

                elif msg_type == "end":
                    # ── Step 0: Save audio locally BEFORE transcription ──
                    await websocket.send_json({
                        "stage": "saving",
                        "message": "Saving audio recording locally…",
                    })
                    recording_meta = {}
                    try:
                        recording_meta = await audio_storage.save_recording(
                            audio_chunks,
                            outgoing_nurse=outgoing_nurse,
                            incoming_nurse=incoming_nurse,
                        )
                        recording_id = recording_meta.get("recording_id", "")
                        if recording_id:
                            await websocket.send_json({
                                "stage": "saved",
                                "message": f"Audio saved locally ({recording_meta.get('original_size_bytes', 0)} bytes)",
                                "recording_id": recording_id,
                            })
                    except Exception as e:
                        print(f"[WS] Audio save warning (non-fatal): {e}")
                        # Non-fatal — continue with transcription even if save fails

                    # ── Step 1: Full transcription (from saved WAV) ──────
                    await websocket.send_json({
                        "stage": "transcribing",
                        "message": "Sending to Python for conversion & transcription…",
                    })
                    try:
                        relay = RelayAgent()
                        for c in audio_chunks:
                            await relay.process_audio_chunk(c)
                        full_transcript = await relay.transcribe_full(language=language)

                        # Save transcript alongside the audio
                        if full_transcript and recording_meta.get("recording_id"):
                            await audio_storage.save_transcript(
                                recording_meta["recording_id"], full_transcript
                            )

                        # Send authoritative final transcript
                        await websocket.send_json({
                            "stage": "transcript",
                            "data": full_transcript,
                            "final": True,
                        })

                        # Do not continue to extraction with fake/demo fallback when real audio failed
                        if not full_transcript or not full_transcript.strip():
                            await websocket.send_json({
                                "stage": "error",
                                "message": "No speech was transcribed from the recording. Please retry and ensure microphone access/input is active.",
                            })
                            break
                    except Exception as e:
                        print(f"[WS] Transcription error: {traceback.format_exc()}")
                        full_transcript = ""
                        await websocket.send_json({
                            "stage": "error",
                            "message": "Transcription failed before SBAR extraction. Please retry.",
                        })
                        break

                    # ── Step 2: Extract → Sentinel → Bridge ──────────────
                    await websocket.send_json({
                        "stage": "extracting",
                        "message": "Extracting SBAR data…",
                    })
                    try:
                        final = await pipeline.run_from_transcript(
                            full_transcript, outgoing_nurse, incoming_nurse,
                        )
                        session_id = await db.save_session(final)
                        final.session_id = session_id

                        # Link the recording to the session
                        if recording_meta.get("recording_id"):
                            meta = await audio_storage.get_recording_metadata(recording_meta["recording_id"])
                            if meta:
                                meta["session_id"] = session_id
                                from pathlib import Path
                                meta_path = Path(audio_storage._METADATA_DIR) / f"{recording_meta['recording_id']}.json"
                                meta_path.write_text(json.dumps(meta, indent=2))

                        await _stream_final(websocket, final)
                    except Exception as e:
                        print(f"[WS] Pipeline error: {traceback.format_exc()}")
                        await websocket.send_json({"stage": "error", "message": str(e)})
                    break

                elif msg_type == "demo":
                    outgoing_nurse = data.get("outgoing", "Nurse Sarah Chen")
                    incoming_nurse = data.get("incoming", "Nurse Marcus Rivera")
                    await websocket.send_json({
                        "stage": "extracting",
                        "message": "Running demo pipeline…",
                    })
                    try:
                        final = await pipeline.run_demo(outgoing_nurse, incoming_nurse)
                        session_id = await db.save_session(final)
                        final.session_id = session_id
                        await _stream_final(websocket, final)
                    except Exception as e:
                        print(f"[WS] Demo pipeline error: {traceback.format_exc()}")
                        await websocket.send_json({"stage": "error", "message": str(e)})
                    break

            elif "bytes" in message:
                chunk = message["bytes"]
                if chunk:
                    audio_chunks.append(chunk)

                    # Live partial transcription every N chunks
                    chunks_since = len(audio_chunks) - last_transcribed_idx
                    if chunks_since >= _LIVE_TRANSCRIBE_EVERY:
                        # Always include chunk 0 (WebM header) for decodable audio
                        if last_transcribed_idx > 0:
                            segment_audio = audio_chunks[0] + b"".join(audio_chunks[last_transcribed_idx:])
                        else:
                            segment_audio = b"".join(audio_chunks)
                        partial = await transcribe_buffer(segment_audio, language=language)
                        if partial:
                            await websocket.send_json({
                                "stage": "transcript",
                                "data": partial,
                            })
                            last_transcribed_idx = len(audio_chunks)

                            # Auto-save draft transcript (non-blocking)
                            try:
                                draft_id = f"live_{outgoing_nurse}_{incoming_nurse}".replace(" ", "_")[:60]
                                await audio_storage.save_draft_transcript(draft_id, partial)
                            except Exception:
                                pass  # non-critical

    except WebSocketDisconnect:
        print("Client disconnected from WebSocket")
    except Exception as e:
        print(f"WebSocket error: {traceback.format_exc()}")
        try:
            await websocket.send_json({"stage": "error", "message": str(e)})
        except Exception:
            pass


async def _stream_final(websocket: WebSocket, final) -> None:
    """Stream each pipeline stage result back to the client."""
    await websocket.send_json({"stage": "extract",   "data": final.sbar.model_dump()})
    await websocket.send_json({"stage": "sentinel",  "data": [a.model_dump() for a in final.alerts]})
    await websocket.send_json({"stage": "bridge",    "data": {"rendered": final.rendered}})
    await websocket.send_json({"stage": "complete",  "data": final.model_dump()})
