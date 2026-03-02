from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "demo" / "medrelay_feed_data.xlsx"

wb = Workbook()

# Remove default sheet
wb.remove(wb.active)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
center = Alignment(horizontal="center", vertical="center")


def style_headers(ws, header_row=1):
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 48)


# Sheet 1: Feed_Instructions
ws = wb.create_sheet("Feed_Instructions")
ws.append(["Field", "Description", "Required", "Example"])
instructions = [
    ("session_id", "Unique handoff ID", "Yes", "HR-2026-03-001"),
    ("timestamp", "Handoff timestamp", "Yes", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ("outgoing_nurse", "Nurse handing off", "Yes", "Nurse Sarah Chen"),
    ("incoming_nurse", "Nurse taking over", "Yes", "Nurse Marcus Rivera"),
    ("patient_name", "Patient full name", "Yes", "Sarah Mitchell"),
    ("patient_age", "Age", "Yes", "67"),
    ("patient_mrn", "Medical record number", "Yes", "ICU-2024-0447"),
    ("patient_room", "Room/bed", "Yes", "ICU 4B"),
    ("primary_diagnosis", "Main diagnosis", "Yes", "Septic shock secondary to pneumonia"),
    ("reason_for_admission", "Reason admitted", "Yes", "Hypotension and hypoxia from severe pneumonia"),
    ("current_status", "Current condition", "Yes", "Hemodynamically unstable on vasopressors"),
    ("vitals_bp", "Blood pressure", "Yes", "88/54"),
    ("vitals_hr", "Heart rate", "Yes", "118"),
    ("vitals_rr", "Respiratory rate", "Yes", "24"),
    ("vitals_temp", "Temperature celsius", "Yes", "38.9"),
    ("vitals_spo2", "SpO2 percentage", "Yes", "91"),
    ("medications", "Pipe-separated medication list", "Yes", "Norepinephrine|Vancomycin|Piperacillin-Tazobactam"),
    ("allergies", "Pipe-separated allergies", "Yes", "Penicillin (anaphylaxis)|Latex"),
    ("labs_pending", "Pipe-separated pending labs", "No", "Blood cultures x2|Repeat lactate"),
    ("labs_recent", "Pipe-separated recent labs", "No", "Lactate 4.2|WBC 18.4"),
    ("care_plan", "Care plan", "Yes", "Continue sepsis protocol and titrate norepinephrine"),
    ("escalation_triggers", "Escalation criteria", "Yes", "MAP < 65 or SpO2 < 88"),
    ("pending_orders", "Pipe-separated orders", "No", "Echo|ID consult"),
    ("next_steps", "Shift handoff next actions", "Yes", "Q1h vitals and family update"),
]
for row in instructions:
    ws.append(list(row))
style_headers(ws)
autosize(ws)


# Sheet 2: Patient_Feed_Template
ws = wb.create_sheet("Patient_Feed_Template")
ws.append([
    "session_id", "timestamp", "outgoing_nurse", "incoming_nurse", "patient_name", "patient_age", "patient_mrn", "patient_room",
    "primary_diagnosis", "reason_for_admission", "current_status", "vitals_bp", "vitals_hr", "vitals_rr", "vitals_temp", "vitals_spo2",
    "medications", "allergies", "labs_pending", "labs_recent", "care_plan", "escalation_triggers", "pending_orders", "next_steps"
])
ws.append([
    "HR-2026-03-001", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Nurse Sarah Chen", "Nurse Marcus Rivera", "Sarah Mitchell", "67", "ICU-2024-0447", "ICU 4B",
    "Septic shock secondary to pneumonia", "Hypotension and hypoxia from severe pneumonia", "Hemodynamically unstable on vasopressor support",
    "88/54", 118, 24, 38.9, 91,
    "Norepinephrine 0.1 mcg/kg/min IV|Vancomycin 1g IV q12h|Piperacillin-Tazobactam 3.375g IV q6h|Heparin 5000 units SubQ q8h",
    "Penicillin (anaphylaxis)|Latex (rash)",
    "Blood cultures x2|Repeat lactate 2h|CBC|BMP",
    "Lactate 4.2 mmol/L|WBC 18.4 x10^9/L|Procalcitonin 22.1 ng/mL",
    "Continue sepsis bundle and titrate norepinephrine to maintain MAP >= 65",
    "MAP < 65|SpO2 < 88|urine output < 30 mL/hr|worsening mentation",
    "Repeat lactate|Echocardiogram|Infectious Disease consult",
    "Q1h vitals, strict I/O, family update at 0800"
])
style_headers(ws)
autosize(ws)


# Sheet 3: Risk_Alerts_Seed
ws = wb.create_sheet("Risk_Alerts_Seed")
ws.append(["session_id", "severity", "category", "description"])
alerts = [
    ("HR-2026-03-001", "HIGH", "medication", "Piperacillin-Tazobactam includes penicillin in patient with anaphylaxis allergy"),
    ("HR-2026-03-001", "HIGH", "vital", "SpO2 91% on high-flow oxygen indicates persistent hypoxemia"),
    ("HR-2026-03-001", "HIGH", "vital", "BP 88/54 indicates ongoing shock despite vasopressor support"),
    ("HR-2026-03-001", "MEDIUM", "vital", "Temperature 38.9°C suggests active infection"),
]
for row in alerts:
    ws.append(list(row))
style_headers(ws)
autosize(ws)


# Sheet 4: Transcript_Seed
ws = wb.create_sheet("Transcript_Seed")
ws.append(["session_id", "transcript"])
ws.append([
    "HR-2026-03-001",
    "Patient Sarah Mitchell in ICU 4B, 67-year-old admitted with septic shock from pneumonia. On norepinephrine, BP 88/54, HR 118, RR 24, temp 38.9, SpO2 91 on high-flow. Allergic to penicillin. Awaiting blood cultures and repeat lactate. Escalate for MAP below 65 or SpO2 below 88."
])
style_headers(ws)
autosize(ws)


OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(str(OUT))
